"""📤 Прочитане зі справи → таблиця, придатна до Ексселю.

Чотири вигляди на ті самі записи, бо питання до них різні:

* **acts** — рядок = АКТ, ролі розкладені в колонки. Так акт читається одним
  поглядом: хто народився, чиї батьки, хто хрестив.
* **records** — рядок = УЧАСНИК. Тут прізвище, стан, вік і місце стоять
  окремими полями, тож саме цей вигляд фільтрується («усі однодворці»,
  «усі, кому за 60»).
* **pages** — рядок = СТОРІНКА: що на ній і чи повний перелік прізвищ.
* **tally** — власні підсумки книги («родилось мужеска 5, женска 4»).

🔴 Підсумки НЕ змішуються з актами. `tally` — це чексум повноти вичитки, а не
подія: потрапивши в один список з актами, він виглядає як ще один акт без
учасників і псує будь-який підрахунок по книзі.

🔴 Колонка зі сканами є в кожному вигляді. Виписка без посилання на аркуш —
переказ: перевірити її можна тільки перечитавши всю справу, тобто ніяк.

Порожнє поле лишається порожнім. Тут нічого не добудовується й не виводиться
«за структурою»: єдине похідне значення — `year`, і воно береться з дати
самого акту, а не з сусідніх аркушів.
"""
from __future__ import annotations

import csv
import re
from collections.abc import Sequence
from pathlib import Path
from typing import Any, Literal

View = Literal["acts", "records", "pages", "tally"]
VIEWS: tuple[View, ...] = ("acts", "records", "pages", "tally")

Fmt = Literal["csv", "tsv", "xlsx"]
FORMATS: tuple[Fmt, ...] = ("csv", "tsv", "xlsx")

#: Роздільник за форматом; xlsx свого не має.
SEPARATOR: dict[str, str] = {"csv": ",", "tsv": "\t"}

# Порядок ролей у широкому вигляді — за ходом самого акту, а не за абеткою:
# спершу той, про кого запис, далі батьки, далі хресні й свідки. Причт
# останній, бо повторюється під кожним актом однаково.
ROLE_ORDER: tuple[str, ...] = (
    "child", "father", "mother", "godfather", "godmother",
    "groom", "groom_father", "groom_mother",
    "bride", "bride_father", "bride_mother", "witness",
    "deceased", "spouse",
    "head", "member",
    "convert", "sponsor",
    "midwife", "priest", "other",
)

ROLE_LABEL: dict[str, str] = {
    "child": "дитина", "father": "батько", "mother": "мати",
    "godfather": "хрещений", "godmother": "хрещена",
    "groom": "наречений", "bride": "наречена",
    "groom_father": "батько нареченого", "groom_mother": "мати нареченого",
    "bride_father": "батько нареченої", "bride_mother": "мати нареченої",
    "deceased": "померлий", "spouse": "подружжя", "witness": "свідок",
    "priest": "причт", "midwife": "повитуха",
    "head": "голова двору", "member": "член двору",
    "convert": "приєднаний", "sponsor": "поручитель", "other": "інше",
}

LABEL: dict[str, str] = {
    "rid": "ключ запису", "type": "тип", "date": "дата події",
    "date2": "дата обряду", "year": "рік", "sheet": "аркуш", "row": "№",
    "scans": "скани", "places": "місця акту", "cause": "причина смерті",
    "quote": "цитата", "confidence": "певність", "comment": "коментар",
    "role": "роль", "name": "як у джерелі", "surname": "прізвище",
    "given": "імʼя", "patronymic": "по батькові", "sex": "стать",
    "estate": "стан", "age": "вік", "place": "місце", "note": "примітка",
    "scan": "скан", "status": "статус", "surnames": "прізвища",
    "years": "роки", "method": "спосіб",
}

#: Той самий ключ у різних виглядах означає різне: у списку актів `places` —
#: місця, названі В АКТІ, а у списку сторінок — усе, що трапилось на аркуші.
#: Спільний підпис зробив би з другого перше, і це виглядало б як дані.
VIEW_LABEL: dict[str, dict[str, str]] = {
    "pages": {"places": "місця", "comment": "що на аркуші"},
}

#: Підписи лічильників книги. «разом» тут — той, що НАПИСАНИЙ у книзі.
_COUNT_LABEL: dict[str, str] = {
    "m": "чоловіча", "f": "жіноча",
    "total": "разом (у книзі)", "total_both_sexes": "разом обох статей",
}

SHEET_TITLE: dict[str, str] = {
    "acts": "Акти", "records": "Учасники",
    "pages": "Сторінки", "tally": "Підсумки",
}

_ACT_HEAD = ("rid", "type", "year", "date", "date2", "sheet", "row", "scans")
_ACT_TAIL = ("places", "cause", "quote", "confidence", "comment")


def label_for(key: str, view: str = "") -> str:
    """Українська шапка колонки; для ролі — назва ролі."""
    if view and key in VIEW_LABEL.get(view, {}):
        return VIEW_LABEL[view][key]
    if key in LABEL:
        return LABEL[key]
    if key in ROLE_LABEL:
        return ROLE_LABEL[key]
    if key.startswith("counts."):
        # Ключі лічильників приходять із самої книги, тож перелік відкритий:
        # знайомі підписуються, незнайомий лишається як записано — вигадана
        # назва графи гірша за англійський ключ.
        return _COUNT_LABEL.get(key[7:], key[7:])
    return key


# ── допоміжне ────────────────────────────────────────────────────────────────
def _dt(value: Any) -> str:
    """GedDate → рядок як записано; None → порожньо."""
    return getattr(value, "value", "") or ""


def _year(rec: Any) -> str:
    """Рік акту з ЙОГО ВЛАСНОЇ дати.

    🔴 Не з сусіднього аркуша: у метричних справах часто підшито два примірники
    блоками, тож сусід може бути з іншого року, і виведений так рік виглядав би
    прочитаним.
    """
    for raw in (_dt(rec.date), _dt(rec.date2)):
        if len(raw) >= 4 and raw[:4].isdigit():
            return raw[:4]
    return ""


_ROW_SPLIT = re.compile(r"(\d+)")


def _sort_key(rec: Any) -> tuple[str, str, int, str]:
    """Порядок аркушів і номерів — щоб таблиця горталась, як сама книга.

    Номер акту не число: у метриках лічильники окремі за статтю («м38», «ж36»).
    Тому ключ — літерний префікс окремо, цифри окремо; без цього «м10» стає
    перед «м9».
    """
    scan = rec.scans[0] if rec.scans else ""
    row = (rec.row or "").strip()
    parts = _ROW_SPLIT.split(row, maxsplit=1)
    prefix = parts[0] if parts else row
    number = int(parts[1]) if len(parts) > 1 else -1
    return (scan, prefix, number, row)


# ── вигляди ──────────────────────────────────────────────────────────────────
def _acts(cf: Any) -> tuple[list[str], list[dict[str, str]]]:
    """Рядок = акт, ролі в колонки.

    Колонки ролей — лише ті, що в цій справі СПРАВДІ трапились: у книзі самих
    народжень порожні «наречена» й «поручитель» тільки заважають фільтрувати.
    """
    acts = sorted((r for r in cf.records if r.rtype != "tally"), key=_sort_key)
    seen: set[str] = {p.role for r in acts for p in r.persons}
    roles = [r for r in ROLE_ORDER if r in seen]
    roles += sorted(seen - set(ROLE_ORDER))
    columns = [*_ACT_HEAD, *roles, *_ACT_TAIL]

    rows: list[dict[str, str]] = []
    for rec in acts:
        by_role: dict[str, list[str]] = {}
        for p in rec.persons:
            # Роль без імені не мовчить: у книзі трапляється «восприемница
            # [нрзб]», і таку появу треба рахувати, а не втрачати.
            by_role.setdefault(p.role, []).append(p.name or "[нрзб]")
        row = {
            "rid": rec.rid, "type": rec.rtype, "year": _year(rec),
            "date": _dt(rec.date), "date2": _dt(rec.date2),
            "sheet": rec.sheet, "row": rec.row, "scans": "; ".join(rec.scans),
            "places": "; ".join(rec.places), "cause": rec.cause or "",
            "quote": rec.quote or "", "confidence": str(rec.confidence),
            "comment": rec.comment,
        }
        # Кілька носіїв однієї ролі (двоє поручителів) склеюються в одну
        # колонку. Окремі «свідок 1 / свідок 2» розсунули б схему на
        # найширшому акті справи, і в решті рядків стояла б порожнеча.
        for role in roles:
            row[role] = "; ".join(by_role.get(role, []))
        rows.append(row)
    return columns, rows


def _records(cf: Any) -> tuple[list[str], list[dict[str, str]]]:
    """Рядок = учасник акту: саме тут фільтрується прізвище, стан і вік."""
    columns = ["rid", "type", "year", "date", "date2", "scans", "sheet", "row",
               "role", "name", "surname", "given", "patronymic", "sex",
               "estate", "age", "place", "note",
               "places", "cause", "quote", "confidence", "comment"]
    rows: list[dict[str, str]] = []
    for rec in sorted((r for r in cf.records if r.rtype != "tally"), key=_sort_key):
        shared = {
            "rid": rec.rid, "type": rec.rtype, "year": _year(rec),
            "date": _dt(rec.date), "date2": _dt(rec.date2),
            "scans": "; ".join(rec.scans), "sheet": rec.sheet, "row": rec.row,
            "places": "; ".join(rec.places), "cause": rec.cause or "",
            "quote": rec.quote or "", "confidence": str(rec.confidence),
            "comment": rec.comment,
        }
        # Акт без жодного розібраного учасника все одно потрапляє в таблицю:
        # інакше сторінка виглядала б невичитаною там, де її читали.
        if not rec.persons:
            rows.append({**shared, "role": "", "name": "", "surname": "",
                         "given": "", "patronymic": "", "sex": "",
                         "estate": "", "age": "", "place": "", "note": ""})
            continue
        for p in rec.persons:
            rows.append({**shared, "role": p.role, "name": p.name,
                         "surname": p.surname or "", "given": p.given or "",
                         "patronymic": p.patronymic or "", "sex": p.sex or "",
                         "estate": p.estate or "", "age": p.age or "",
                         "place": p.place or "", "note": p.note or ""})
    return columns, rows


def _pages(cf: Any) -> tuple[list[str], list[dict[str, str]]]:
    columns = ["scan", "type", "status", "sheet", "years",
               "surnames", "places", "method", "comment"]
    rows = [{"scan": n.scan, "type": n.page_type, "status": n.status,
             "sheet": n.sheet, "years": "; ".join(str(y) for y in n.years),
             "surnames": "; ".join(n.surnames), "places": "; ".join(n.places),
             "method": n.method, "comment": n.comment}
            for n in sorted(cf.pages.values(), key=lambda n: n.scan)]
    return columns, rows


def _tally(cf: Any) -> tuple[list[str], list[dict[str, str]]]:
    """Власні підсумки книги — окремо від актів, бо це чексум, а не подія."""
    recs = sorted((r for r in cf.records if r.rtype == "tally"), key=_sort_key)
    keys: list[str] = []
    for rec in recs:
        for k in rec.counts:
            if k not in keys:
                keys.append(k)
    keys.sort(key=lambda k: ({"m": 0, "f": 1}.get(k, 2), k))
    # 🔴 Колонки «разом» тут НЕМАЄ, і це не забуто. Підсумок належить книзі: у
    # цій самій справі поряд із «мужеска/женска» трапляються власні `total` і
    # `total_both_sexes`, тож сума по всіх ключах порахувала б їх удруге — і
    # вийшло б обчислене число, яке виглядає як прочитане. Чексум звіряє
    # `audit`, зіставляючи підсумок книги з нумерацією актів.
    columns = ["rid", "year", "date", "scans", "sheet", "row",
               *(f"counts.{k}" for k in keys), "comment"]
    rows: list[dict[str, str]] = []
    for rec in recs:
        row = {"rid": rec.rid, "year": _year(rec), "date": _dt(rec.date),
               "scans": "; ".join(rec.scans), "sheet": rec.sheet,
               "row": rec.row, "comment": rec.comment}
        for k in keys:
            row[f"counts.{k}"] = str(rec.counts[k]) if k in rec.counts else ""
        rows.append(row)
    return columns, rows


_BUILDERS = {"acts": _acts, "records": _records, "pages": _pages, "tally": _tally}


def build(cf: Any, what: str) -> tuple[list[str], list[dict[str, str]]]:
    """Вигляд `what` для справи: (колонки, рядки)."""
    builder = _BUILDERS.get(what)
    if builder is None:
        raise ValueError(f"невідомий вигляд «{what}»; є: {', '.join(VIEWS)}")
    return builder(cf)


# ── запис у файл ─────────────────────────────────────────────────────────────
# 🔴 Кирилиця в Екселі живе на BOM. Без нього виписка відкривається
# «крякозябрами», і виглядає це як зіпсовані ДАНІ, а не як кодування — людина
# вирішує, що вичитка провалилась. Та сама причина, що й у браузерній кнопці
# CSV (`daemon/static/screens/export.js`).
CSV_ENCODING = "utf-8-sig"

#: Ексель не приймає керівні символи в комірці, а декод скоропису їх приносить.
_ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

#: Стеля довжини комірки в самому форматі XLSX.
_CELL_MAX = 32767


class ExportError(RuntimeError):
    """Записати не вийшло — з причиною, придатною для показу людині."""


def _clean(value: str) -> tuple[str, bool, bool]:
    """Комірка → (значення, чи чистили символи, чи різали довжину)."""
    stripped = _ILLEGAL.sub("", value)
    had_bad = stripped != value
    if len(stripped) > _CELL_MAX:
        return stripped[:_CELL_MAX], had_bad, True
    return stripped, had_bad, False


def _header(columns: list[str], human: bool, view: str = "") -> list[str]:
    return [label_for(c, view) for c in columns] if human else list(columns)


def write_delimited(path: Any, columns: list[str], rows: list[dict[str, str]],
                    *, sep: str = ",", human: bool = True,
                    view: str = "") -> dict[str, Any]:
    """CSV/TSV одного вигляду. Один файл — одна таблиця, це властивість формату."""
    dest = Path(path)
    dest.parent.mkdir(parents=True, exist_ok=True)
    with dest.open("w", encoding=CSV_ENCODING, newline="") as fh:
        writer = csv.writer(fh, delimiter=sep, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(_header(columns, human, view))
        for row in rows:
            writer.writerow([row.get(c, "") for c in columns])
    return {"path": str(dest), "rows": len(rows), "sheets": 1,
            "cleaned": 0, "truncated": 0}


def write_xlsx(path: Any,
               sheets: Sequence[tuple[str, list[str], list[dict[str, str]]]],
               *, human: bool = True) -> dict[str, Any]:
    """Кілька виглядів одним файлом: акти, учасники, сторінки, підсумки."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:  # pragma: no cover — залежить від складу встановлення
        raise ExportError(
            "формат xlsx потребує openpyxl: "
            "pip install 'nyshporka[xlsx]' — або взяти --format csv, "
            "він працює без додаткових пакетів") from None

    dest = Path(path)
    dest.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    # Порожня книга приходить з одним аркушем «Sheet»; свої додаються нижче.
    blank = book.active
    if blank is not None:
        book.remove(blank)
    cleaned = truncated = total = 0

    for view, columns, rows in sheets:
        sheet = book.create_sheet(SHEET_TITLE.get(view, view)[:31])
        sheet.append(_header(columns, human, view))
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in rows:
            values: list[str] = []
            for column in columns:
                value, had_bad, cut = _clean(str(row.get(column, "")))
                cleaned += had_bad
                truncated += cut
                values.append(value)
            sheet.append(values)
            # 🔴 Ексель читає провідний «=» як формулу, і openpyxl цю здогадку
            # повторює. Прізвище чи цитата з таким початком перетворились би на
            # #NAME? — тобто прочитане з аркуша зникло б з таблиці.
            for cell in sheet[sheet.max_row]:
                if isinstance(cell.value, str) and cell.value[:1] in "=+-@":
                    cell.data_type = "s"
        total += len(rows)

        if rows:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = (
                f"A1:{get_column_letter(len(columns))}{len(rows) + 1}")
        for index, column in enumerate(columns, start=1):
            # Ширина за вмістом перших рядків: повний прохід по 30 тисячах
            # комірок коштує більше, ніж дає.
            widest = max([len(_header([column], human, view)[0])]
                         + [len(str(r.get(column, ""))) for r in rows[:200]])
            sheet.column_dimensions[get_column_letter(index)].width = \
                min(max(widest + 2, 8), 60)

    book.save(dest)
    return {"path": str(dest), "rows": total, "sheets": len(sheets),
            "cleaned": cleaned, "truncated": truncated}
